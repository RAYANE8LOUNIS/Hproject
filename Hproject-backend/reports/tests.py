# Student 5 - Reports dashboard and export tests.

from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from organisation.models import Department, Team


class ReportsViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='student5',
            password='testpass123',
            email='student5@example.com',
        )
        self.manager = User.objects.create_user(
            username='manager',
            password='testpass123',
            first_name='Maya',
            last_name='Patel',
        )
        self.members = [
            User.objects.create_user(username=f'engineer{i}', password='testpass123')
            for i in range(5)
        ]
        self.department = Department.objects.create(
            name='Platform Engineering',
            specialisation='Developer Experience',
            leader=self.manager,
        )
        self.managed_team = Team.objects.create(
            name='Portal Core',
            department=self.department,
            manager=self.manager,
            email='portal-core@example.com',
            code_repository='https://example.com/portal-core',
            status='active',
        )
        self.managed_team.members.add(*self.members)
        self.unmanaged_team = Team.objects.create(
            name='Unassigned Integrations',
            department=self.department,
            slack_channel='#unassigned-integrations',
            status='restructured',
        )

    def test_reports_dashboard_requires_login(self):
        response = self.client.get(reverse('reports:dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/accounts/login/', response.url)

    def test_reports_dashboard_displays_summary_and_unmanaged_team(self):
        self.client.login(username='student5', password='testpass123')
        response = self.client.get(reverse('reports:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Engineering Team Reports')
        self.assertContains(response, 'Platform Engineering')
        self.assertContains(response, 'Unassigned Integrations')
        self.assertEqual(response.context['total_teams'], 2)
        self.assertEqual(len(response.context['teams_without_managers']), 1)

    def test_pdf_download_returns_valid_pdf(self):
        self.client.login(username='student5', password='testpass123')
        response = self.client.get(reverse('reports:download_pdf'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF-1.4'))
        self.assertIn(b'Unassigned Integrations', response.content)

    def test_excel_download_returns_report_workbook(self):
        self.client.login(username='student5', password='testpass123')
        response = self.client.get(reverse('reports:download_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertIn('Summary', workbook.sheetnames)
        self.assertIn('Department Summary', workbook.sheetnames)
        self.assertIn('Teams Without Managers', workbook.sheetnames)
        self.assertEqual(workbook['Summary']['A3'].value, 'Total teams')
        self.assertEqual(workbook['Summary']['B3'].value, 2)
        self.assertEqual(workbook['Teams Without Managers']['A2'].value, 'Unassigned Integrations')
