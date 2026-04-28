# Student 5 - Reports (PDF/Excel) implementation.

from io import BytesIO
from textwrap import wrap

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from organisation.models import Department, Dependency, Team


def _display_user(user):
    if not user:
        return 'No manager assigned'
    return user.get_full_name() or user.username


def _team_contact(team):
    contacts = []
    if team.email:
        contacts.append(team.email)
    if team.slack_channel:
        contacts.append(team.slack_channel)
    return ', '.join(contacts) if contacts else 'No contact channel'


def _status_summaries(teams):
    return [
        {
            'key': key,
            'label': label,
            'count': sum(1 for team in teams if team.status == key),
        }
        for key, label in Team.STATUS_CHOICES
    ]


def _department_summaries(departments, teams):
    summaries = []
    for department in departments:
        department_teams = [team for team in teams if team.department_id == department.id]
        summaries.append({
            'department': department,
            'team_count': len(department_teams),
            'manager_count': len({team.manager_id for team in department_teams if team.manager_id}),
            'member_count': sum(team.members.count() for team in department_teams),
            'repository_count': sum(1 for team in department_teams if team.code_repository),
            'without_manager_count': sum(1 for team in department_teams if not team.manager_id),
            'active_count': sum(1 for team in department_teams if team.status == 'active'),
        })
    return summaries


def _build_report_data():
    teams = list(
        Team.objects.select_related('department', 'manager', 'team_type')
        .prefetch_related('members')
        .order_by('name')
    )
    departments = list(
        Department.objects.select_related('leader')
        .prefetch_related('teams')
        .order_by('name')
    )
    dependencies = list(
        Dependency.objects.select_related(
            'upstream_team',
            'upstream_team__department',
            'downstream_team',
            'downstream_team__department',
        ).order_by('upstream_team__name', 'downstream_team__name')
    )

    teams_without_managers = [team for team in teams if not team.manager_id]
    teams_below_member_requirement = [team for team in teams if team.members.count() < 5]
    teams_without_repository = [team for team in teams if not team.code_repository]
    teams_without_contact = [
        team for team in teams
        if not team.email and not team.slack_channel
    ]
    departments_below_team_requirement = [
        summary for summary in _department_summaries(departments, teams)
        if summary['team_count'] < 3
    ]

    department_summaries = _department_summaries(departments, teams)
    member_ids = set()
    for team in teams:
        member_ids.update(team.members.values_list('id', flat=True))

    return {
        'generated_at': timezone.now(),
        'teams': teams,
        'departments': departments,
        'dependencies': dependencies,
        'department_summaries': department_summaries,
        'status_summaries': _status_summaries(teams),
        'teams_without_managers': teams_without_managers,
        'teams_below_member_requirement': teams_below_member_requirement,
        'departments_below_team_requirement': departments_below_team_requirement,
        'teams_without_repository': teams_without_repository,
        'teams_without_contact': teams_without_contact,
        'total_teams': len(teams),
        'total_departments': len(departments),
        'total_managers': len({team.manager_id for team in teams if team.manager_id}),
        'total_members': len(member_ids),
        'total_repositories': sum(1 for team in teams if team.code_repository),
        'total_dependencies': len(dependencies),
    }


@login_required
def dashboard(request):
    return render(request, 'reports/reports_dashboard.html', _build_report_data())


def _escape_pdf_text(value):
    return str(value).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _pdf_lines(data):
    lines = [
        'Sky Engineering Team Portal - Reports',
        f"Generated: {timezone.localtime(data['generated_at']).strftime('%d %b %Y %H:%M')}",
        '',
        'Summary',
        f"Total teams: {data['total_teams']}",
        f"Departments: {data['total_departments']}",
        f"Managers assigned: {data['total_managers']}",
        f"Distinct team members: {data['total_members']}",
        f"Repositories recorded: {data['total_repositories']}",
        f"Dependencies recorded: {data['total_dependencies']}",
        f"Teams without managers: {len(data['teams_without_managers'])}",
        '',
        'Status breakdown',
    ]

    for item in data['status_summaries']:
        lines.append(f"{item['label']}: {item['count']}")

    lines.extend(['', 'Department summary'])
    for summary in data['department_summaries']:
        department = summary['department']
        lines.append(
            f"{department.name}: {summary['team_count']} teams, "
            f"{summary['member_count']} members, "
            f"{summary['repository_count']} repositories, "
            f"{summary['without_manager_count']} without manager"
        )

    lines.extend(['', 'Teams without managers'])
    if data['teams_without_managers']:
        for team in data['teams_without_managers']:
            lines.append(
                f"{team.name} | Department: {team.department.name if team.department else 'Not assigned'} | "
                f"Contact: {_team_contact(team)}"
            )
    else:
        lines.append('No teams without managers were found in the current database.')

    lines.extend(['', 'Data quality checks'])
    lines.append(f"Teams below five members: {len(data['teams_below_member_requirement'])}")
    lines.append(f"Departments below three teams: {len(data['departments_below_team_requirement'])}")
    lines.append(f"Teams without repository: {len(data['teams_without_repository'])}")
    lines.append(f"Teams without contact channel: {len(data['teams_without_contact'])}")
    return lines


def _build_simple_pdf(lines):
    printable_lines = []
    for line in lines:
        wrapped = wrap(str(line), width=96, replace_whitespace=False) or ['']
        printable_lines.extend(wrapped)

    page_chunks = [
        printable_lines[index:index + 45]
        for index in range(0, len(printable_lines), 45)
    ] or [['No report data available.']]

    objects = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        None,
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
    ]
    page_refs = []

    for chunk in page_chunks:
        page_number = len(objects) + 1
        content_number = page_number + 1
        page_refs.append(f'{page_number} 0 R')

        content_lines = ['BT', '/F1 10 Tf', '50 792 Td', '14 TL']
        for line in chunk:
            content_lines.append(f'({_escape_pdf_text(line)}) Tj')
            content_lines.append('T*')
        content_lines.append('ET')
        content = '\n'.join(content_lines).encode('latin-1', errors='replace')
        stream = (
            f'<< /Length {len(content)} >>\nstream\n'.encode('ascii')
            + content
            + b'\nendstream'
        )

        objects.append(
            f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] '
            f'/Resources << /Font << /F1 3 0 R >> >> /Contents {content_number} 0 R >>'
            .encode('ascii')
        )
        objects.append(stream)

    objects[1] = (
        f"<< /Type /Pages /Kids [{' '.join(page_refs)}] /Count {len(page_refs)} >>"
        .encode('ascii')
    )

    output = b'%PDF-1.4\n'
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output += f'{index} 0 obj\n'.encode('ascii') + obj + b'\nendobj\n'

    xref_position = len(output)
    output += f'xref\n0 {len(objects) + 1}\n'.encode('ascii')
    output += b'0000000000 65535 f \n'
    for offset in offsets[1:]:
        output += f'{offset:010d} 00000 n \n'.encode('ascii')
    output += (
        f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n'
        f'startxref\n{xref_position}\n%%EOF'
    ).encode('ascii')
    return output


@login_required
def download_pdf(request):
    data = _build_report_data()
    pdf = _build_simple_pdf(_pdf_lines(data))
    filename = f"sky-team-report-{timezone.localdate():%Y%m%d}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_worksheet(worksheet):
    header_fill = PatternFill('solid', fgColor='122476')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center')

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        width = min(max(max_length + 3, 12), 42)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    worksheet.freeze_panes = 'A2'


@login_required
def download_excel(request):
    data = _build_report_data()
    workbook = Workbook()

    summary = workbook.active
    summary.title = 'Summary'
    summary.append(['Metric', 'Value'])
    summary.append(['Generated at', timezone.localtime(data['generated_at']).strftime('%d %b %Y %H:%M')])
    summary.append(['Total teams', data['total_teams']])
    summary.append(['Departments', data['total_departments']])
    summary.append(['Managers assigned', data['total_managers']])
    summary.append(['Distinct team members', data['total_members']])
    summary.append(['Repositories recorded', data['total_repositories']])
    summary.append(['Dependencies recorded', data['total_dependencies']])
    summary.append(['Teams without managers', len(data['teams_without_managers'])])
    _style_worksheet(summary)

    departments = workbook.create_sheet('Department Summary')
    departments.append([
        'Department', 'Leader', 'Specialisation', 'Teams', 'Active Teams',
        'Managers', 'Members', 'Repositories', 'Teams Without Managers'
    ])
    for item in data['department_summaries']:
        department = item['department']
        departments.append([
            department.name,
            _display_user(department.leader),
            department.specialisation or 'Not specified',
            item['team_count'],
            item['active_count'],
            item['manager_count'],
            item['member_count'],
            item['repository_count'],
            item['without_manager_count'],
        ])
    _style_worksheet(departments)

    unmanaged = workbook.create_sheet('Teams Without Managers')
    unmanaged.append(['Team', 'Department', 'Status', 'Members', 'Repository', 'Contact'])
    for team in data['teams_without_managers']:
        unmanaged.append([
            team.name,
            team.department.name if team.department else 'Not assigned',
            team.get_status_display(),
            team.members.count(),
            team.code_repository or 'No repository',
            _team_contact(team),
        ])
    if not data['teams_without_managers']:
        unmanaged.append(['No teams without managers were found.', '', '', '', '', ''])
    _style_worksheet(unmanaged)

    quality = workbook.create_sheet('Data Quality')
    quality.append(['Check', 'Count', 'Coursework Expectation'])
    quality.append(['Teams below five members', len(data['teams_below_member_requirement']), 'Each team should have at least 5 engineers'])
    quality.append(['Departments below three teams', len(data['departments_below_team_requirement']), 'Each department should have at least 3 teams'])
    quality.append(['Teams without repository', len(data['teams_without_repository']), 'Teams should record code repositories'])
    quality.append(['Teams without contact channel', len(data['teams_without_contact']), 'Teams should record contact channels'])
    _style_worksheet(quality)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"sky-team-report-{timezone.localdate():%Y%m%d}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
